"""
AWS Lambda: Process Pyramid Analytics email reports from S3.

Trigger: S3 PutObject on store-margin-tool-data/incoming/*
Flow:    Parse email → extract Excel attachment → identify report type & store
         → parse rows → upsert to RDS → move file to archive folder.
"""

import email
import io
import json
import logging
import os
import re
from datetime import datetime, timezone
from urllib.parse import unquote_plus

import boto3
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

logger = logging.getLogger()
logger.setLevel(logging.INFO)

S3_BUCKET = os.environ.get("S3_BUCKET", "store-margin-tool-data")

DB_HOST = os.environ.get("DB_HOST", "store-margin-tool.cncmsw0k8ivw.eu-north-1.rds.amazonaws.com")
DB_PORT = int(os.environ.get("DB_PORT", "5432"))
DB_NAME = os.environ.get("DB_NAME", "postgres")
DB_USER = os.environ.get("DB_USER", "postgres")
DB_PASS = os.environ.get("DB_PASS", "DromoreWestFeeney2003")

s3 = boto3.client("s3")

# ── Report type detection ────────────────────────────────────────────

# Ordered list — longer/more-specific patterns first to avoid false matches
# (e.g. "margin report" must match top-sellers before "margin" matches margin-report)
REPORT_PATTERNS: list[tuple[str, list[str]]] = [
    ("fh-coffee", [
        "f&h", "f_h", "frank & honest", "frank_honest", "frank-honest",
        "fh coffee", "fh_coffee", "fh-coffee", "coffee report",
    ]),
    ("top-sellers", [
        "retailer margin report", "retailer_margin_report", "retailer-margin-report",
        "underlying margin", "margin report", "margin_report",
        "top-sellers", "top_sellers", "topsellers",
        "retailer-margin", "retailer_margin",
    ]),
    ("margin-report",  ["margin-report", "marginreport"]),
    ("departments",    ["department", "departments", "dept"]),
    ("cpu-comparisons",["cpu-comparison", "cpu_comparison", "cpucomparison", "cpu"]),
]


def detect_report_type(filename: str) -> str | None:
    """Match filename against known report type keywords (longest match first)."""
    lower = filename.lower()
    for report_type, keywords in REPORT_PATTERNS:
        for kw in keywords:
            if kw in lower:
                return report_type
    return None


def detect_store_number(filename: str) -> str | None:
    """Extract store number (2064 or 2056) from filename."""
    match = re.search(r"(2064|2056)", filename)
    return match.group(1) if match else None


# ── Email parsing ────────────────────────────────────────────────────

def extract_attachment(raw_email: bytes) -> tuple[str, bytes] | None:
    """
    Parse a raw email (as saved by SES) and return the first
    Excel/JSON attachment as (filename, content_bytes).
    """
    msg = email.message_from_bytes(raw_email)

    for part in msg.walk():
        content_type = part.get_content_type()
        filename = part.get_filename() or ""

        # Match Excel or JSON attachments
        is_excel = (
            content_type in (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
                "application/octet-stream",
            )
            and (filename.endswith(".xlsx") or filename.endswith(".xls"))
        )
        is_json = content_type == "application/json" or filename.endswith(".json")

        if is_excel or is_json:
            payload = part.get_payload(decode=True)
            if payload:
                logger.info("Found attachment: %s (%s, %d bytes)", filename, content_type, len(payload))
                return filename, payload

    # Fallback: if the raw file isn't an email (direct Excel upload), return as-is
    # Check if the raw bytes start with the XLSX magic bytes (PK zip header)
    if raw_email[:2] == b"PK":
        logger.info("Raw file is a direct Excel upload (not an email)")
        return "direct-upload.xlsx", raw_email

    # Check if it looks like JSON
    stripped = raw_email.strip()
    if stripped[:1] in (b"[", b"{"):
        logger.info("Raw file is a direct JSON upload (not an email)")
        return "direct-upload.json", raw_email

    return None


# ── Excel parsing: Retailer Margin / Top Sellers ─────────────────────

def parse_top_sellers(wb_bytes: bytes, store_number: str) -> list[dict]:
    """
    Parse the retailer margin Excel report.
    Data starts at row 6 (1-indexed).
    Columns:
      A=category, B=subcategory, D=product name, E=lv_code
      F=l7d_sales, G=l7d_qty, H=l7d_margin, I=l7d_margin_pct
      J=ly_sales,  K=ly_qty,  L=ly_margin,  M=ly_margin_pct
      N=ytd_sales, O=ytd_qty, P=ytd_margin, Q=ytd_margin_pct
      R=yd_sales,  S=yd_qty,  T=yd_margin,  U=yd_margin_pct
    """
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []

    # Track last-seen category/subcategory (they may be merged / only on first row)
    current_category = ""
    current_subcategory = ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        # Need at least columns A-U (21 columns)
        cells = list(row) + [None] * max(0, 21 - len(row))

        # Category / subcategory: carry forward if blank
        if cells[0]:
            current_category = str(cells[0]).strip()
        if cells[1]:
            current_subcategory = str(cells[1]).strip()

        # Product name in column D (index 3)
        raw_name = cells[3]
        if not raw_name:
            continue  # skip blank rows

        # Strip leading number pattern (e.g. "123 - Product Name" → "Product Name")
        name = re.sub(r"^\d+\s*-\s*", "", str(raw_name).strip())
        if not name:
            continue

        lv_code = str(cells[4]).strip() if cells[4] else ""

        def to_float(v) -> float | None:
            if v is None:
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None

        record = {
            "store_number":   store_number,
            "name":           name,
            "lv_code":        lv_code,
            "category":       current_category,
            "subcategory":    current_subcategory,
            "l7d_sales":      to_float(cells[5]),
            "l7d_qty":        to_float(cells[6]),
            "l7d_margin":     to_float(cells[7]),
            "l7d_margin_pct": to_float(cells[8]),
            "ly_sales":       to_float(cells[9]),
            "ly_qty":         to_float(cells[10]),
            "ly_margin":      to_float(cells[11]),
            "ly_margin_pct":  to_float(cells[12]),
            "ytd_sales":      to_float(cells[13]),
            "ytd_qty":        to_float(cells[14]),
            "ytd_margin":     to_float(cells[15]),
            "ytd_margin_pct": to_float(cells[16]),
            "yd_sales":       to_float(cells[17]),
            "yd_qty":         to_float(cells[18]),
            "yd_margin":      to_float(cells[19]),
            "yd_margin_pct":  to_float(cells[20]),
        }
        rows.append(record)

    wb.close()
    logger.info("Parsed %d product rows from Excel for store %s", len(rows), store_number)
    return rows


# ── Excel parsing: F&H Coffee ─────────────────────────────────────────

def parse_fh_coffee(wb_bytes: bytes, store_number: str) -> list[dict]:
    """
    Parse the F&H Coffee report Excel (5-column-per-period format).
    Data starts at row 6 (1-indexed).
    Columns (5 per period, no separate margin_pct):
      A=name (strip leading number), B=lv_code
      L7D: C=sales, D=qty, E=margin, F=waste_pct, G=waste_cups
      LY:  H=sales, I=qty, J=margin, K=waste_pct, L=waste_cups
      YTD: M=sales, N=qty, O=margin, P=waste_pct, Q=waste_cups
      YD:  R=sales, S=qty, T=margin, U=waste_pct, V=waste_cups
    margin_pct is calculated as margin/sales where sales > 0.
    """
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = []

    def to_float(v) -> float | None:
        if v is None:
            return None
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    def to_int(v) -> int | None:
        if v is None:
            return None
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None

    def calc_pct(margin, sales) -> float | None:
        if margin is not None and sales is not None and sales > 0:
            return margin / sales
        return None

    for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        cells = list(row) + [None] * max(0, 22 - len(row))

        raw_name = cells[0]
        if not raw_name:
            continue

        name = re.sub(r"^\d+\s*-\s*", "", str(raw_name).strip())
        if not name:
            continue

        lv_code = str(cells[1]).strip() if cells[1] else ""

        l7d_s, l7d_q, l7d_m = to_float(cells[2]), to_int(cells[3]), to_float(cells[4])
        ly_s,  ly_q,  ly_m  = to_float(cells[7]), to_int(cells[8]), to_float(cells[9])
        ytd_s, ytd_q, ytd_m = to_float(cells[12]), to_int(cells[13]), to_float(cells[14])
        yd_s,  yd_q,  yd_m  = to_float(cells[17]), to_int(cells[18]), to_float(cells[19])

        record = {
            "store_number":   store_number,
            "name":           name,
            "lv_code":        lv_code,
            "l7d_sales":      l7d_s,
            "l7d_qty":        l7d_q,
            "l7d_margin":     l7d_m,
            "l7d_margin_pct": calc_pct(l7d_m, l7d_s),
            "l7d_waste_pct":  to_float(cells[5]),
            "l7d_waste_cups": to_int(cells[6]),
            "ly_sales":       ly_s,
            "ly_qty":         ly_q,
            "ly_margin":      ly_m,
            "ly_margin_pct":  calc_pct(ly_m, ly_s),
            "ly_waste_pct":   to_float(cells[10]),
            "ly_waste_cups":  to_int(cells[11]),
            "ytd_sales":      ytd_s,
            "ytd_qty":        ytd_q,
            "ytd_margin":     ytd_m,
            "ytd_margin_pct": calc_pct(ytd_m, ytd_s),
            "ytd_waste_pct":  to_float(cells[15]),
            "ytd_waste_cups": to_int(cells[16]),
            "yd_sales":       yd_s,
            "yd_qty":         yd_q,
            "yd_margin":      yd_m,
            "yd_margin_pct":  calc_pct(yd_m, yd_s),
            "yd_waste_pct":   to_float(cells[20]),
            "yd_waste_cups":  to_int(cells[21]),
        }
        rows.append(record)

    wb.close()
    logger.info("Parsed %d F&H coffee rows from Excel for store %s", len(rows), store_number)
    return rows


# ── Database operations ──────────────────────────────────────────────

def get_db_connection():
    return psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASS,
        sslmode="require",
        connect_timeout=10,
    )


def upsert_top_sellers(rows: list[dict], store_number: str):
    """Delete existing rows for this store, then bulk insert."""
    if not rows:
        logger.warning("No rows to insert for store %s", store_number)
        return 0

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            # Clear existing data for this store
            cur.execute("DELETE FROM top_sellers WHERE store_number = %s", (store_number,))
            deleted = cur.rowcount
            logger.info("Deleted %d existing rows for store %s", deleted, store_number)

            # Bulk insert
            columns = [
                "store_number", "name", "lv_code", "category", "subcategory",
                "l7d_sales", "l7d_qty", "l7d_margin", "l7d_margin_pct",
                "ly_sales", "ly_qty", "ly_margin", "ly_margin_pct",
                "ytd_sales", "ytd_qty", "ytd_margin", "ytd_margin_pct",
                "yd_sales", "yd_qty", "yd_margin", "yd_margin_pct",
            ]
            values = [
                tuple(row[col] for col in columns)
                for row in rows
            ]

            insert_sql = f"""
                INSERT INTO top_sellers ({', '.join(columns)})
                VALUES %s
            """
            execute_values(cur, insert_sql, values, page_size=500)
            logger.info("Inserted %d rows for store %s", len(values), store_number)

        conn.commit()
        return len(values)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def upsert_fh_coffee(rows: list[dict], store_number: str):
    """Delete existing rows for this store, then bulk insert."""
    if not rows:
        logger.warning("No F&H coffee rows to insert for store %s", store_number)
        return 0

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM fh_coffee WHERE store_number = %s", (store_number,))
            deleted = cur.rowcount
            logger.info("Deleted %d existing F&H coffee rows for store %s", deleted, store_number)

            columns = [
                "store_number", "name", "lv_code",
                "l7d_sales", "l7d_qty", "l7d_margin", "l7d_margin_pct", "l7d_waste_pct", "l7d_waste_cups",
                "ly_sales", "ly_qty", "ly_margin", "ly_margin_pct", "ly_waste_pct", "ly_waste_cups",
                "ytd_sales", "ytd_qty", "ytd_margin", "ytd_margin_pct", "ytd_waste_pct", "ytd_waste_cups",
                "yd_sales", "yd_qty", "yd_margin", "yd_margin_pct", "yd_waste_pct", "yd_waste_cups",
            ]
            values = [
                tuple(row[col] for col in columns)
                for row in rows
            ]

            insert_sql = f"""
                INSERT INTO fh_coffee ({', '.join(columns)})
                VALUES %s
            """
            execute_values(cur, insert_sql, values, page_size=500)
            logger.info("Inserted %d F&H coffee rows for store %s", len(values), store_number)

        conn.commit()
        return len(values)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ── S3 file management ──────────────────────────────────────────────

def move_s3_file(bucket: str, source_key: str, store_number: str, report_type: str):
    """Move processed file from incoming/ to {store}/{report_type}/."""
    filename = source_key.split("/")[-1]
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    dest_key = f"{store_number}/{report_type}/{timestamp}_{filename}"

    s3.copy_object(
        Bucket=bucket,
        CopySource={"Bucket": bucket, "Key": source_key},
        Key=dest_key,
    )
    s3.delete_object(Bucket=bucket, Key=source_key)
    logger.info("Moved s3://%s/%s → s3://%s/%s", bucket, source_key, bucket, dest_key)
    return dest_key


# ── Lambda handler ───────────────────────────────────────────────────

def lambda_handler(event, context):
    """
    Entry point. Triggered by S3 PutObject on incoming/*.
    """
    logger.info("Event: %s", json.dumps(event, default=str))

    results = []

    for record in event.get("Records", []):
        s3_info = record.get("s3", {})
        bucket = s3_info.get("bucket", {}).get("name", S3_BUCKET)
        key = unquote_plus(s3_info.get("object", {}).get("key", ""))

        if not key.startswith("incoming/"):
            logger.info("Skipping non-incoming key: %s", key)
            continue

        logger.info("Processing s3://%s/%s", bucket, key)

        try:
            # 1. Download the file from S3
            response = s3.get_object(Bucket=bucket, Key=key)
            raw_bytes = response["Body"].read()
            logger.info("Downloaded %d bytes from S3", len(raw_bytes))

            # 2. Extract attachment from email (or handle direct upload)
            attachment = extract_attachment(raw_bytes)
            if not attachment:
                raise ValueError(f"No Excel/JSON attachment found in {key}")

            att_filename, att_bytes = attachment
            logger.info("Attachment: %s (%d bytes)", att_filename, len(att_bytes))

            # 3. Detect report type and store number
            #    Try the S3 key first, then fall back to attachment filename
            report_type = detect_report_type(key) or detect_report_type(att_filename)
            store_number = detect_store_number(key) or detect_store_number(att_filename)

            if not report_type:
                raise ValueError(f"Could not identify report type from '{key}' or '{att_filename}'")
            if not store_number:
                if report_type == "fh-coffee":
                    store_number = "2064"
                    logger.info("F&H coffee report — defaulting to store 2064")
                else:
                    raise ValueError(f"Could not identify store number from '{key}' or '{att_filename}'")

            logger.info("Report type: %s, Store: %s", report_type, store_number)

            # 4. Parse and insert based on report type
            inserted = 0

            if report_type == "top-sellers":
                if att_filename.endswith(".json"):
                    # JSON format: assume array of row objects
                    data = json.loads(att_bytes)
                    rows = []
                    for item in data:
                        item["store_number"] = store_number
                        rows.append(item)
                    inserted = upsert_top_sellers(rows, store_number)
                else:
                    # Excel format
                    rows = parse_top_sellers(att_bytes, store_number)
                    inserted = upsert_top_sellers(rows, store_number)

            elif report_type == "fh-coffee":
                rows = parse_fh_coffee(att_bytes, store_number)
                inserted = upsert_fh_coffee(rows, store_number)

            elif report_type in ("margin-report", "departments", "cpu-comparisons"):
                # Placeholder for other report types
                logger.info("Report type '%s' processing not yet implemented — file archived only", report_type)

            else:
                logger.warning("Unknown report type: %s", report_type)

            # 5. Move file to archive folder
            dest_key = move_s3_file(bucket, key, store_number, report_type)

            result = {
                "status": "success",
                "source_key": key,
                "dest_key": dest_key,
                "report_type": report_type,
                "store_number": store_number,
                "rows_inserted": inserted,
                "attachment_filename": att_filename,
            }
            logger.info("Result: %s", json.dumps(result))
            results.append(result)

        except Exception as e:
            logger.exception("Failed to process %s", key)
            results.append({
                "status": "error",
                "source_key": key,
                "error": str(e),
            })

    status_code = 200 if all(r["status"] == "success" for r in results) else 500

    return {
        "statusCode": status_code,
        "body": json.dumps(results, default=str),
    }
